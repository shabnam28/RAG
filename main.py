import azure.functions as func
import logging
import io
import os
from typing import List, Dict, Optional
from dataclasses import dataclass
import PyPDF2
from pathlib import Path
import json
from openai import OpenAI
from openai import AzureOpenAI
from azure.storage.blob import BlobServiceClient
from urllib.parse import urlparse, unquote
import pyodbc
from openpyxl import load_workbook
from io import BytesIO

# =========================
# EMBEDDING CLIENT (Gemini example)
# =========================

PROJECT_ID = os.getenv("PROJECT_ID")
LOCATION = os.getenv("LOCATION")
EMBEDDING_MODEL= os.getenv("EMBEDDING_MODEL")

# =========================
# CV PROCESSOR
# =========================

class AzureSQLClient:
    def __init__(self):
        self.server = os.getenv("AZURE_SQL_SERVER")
        self.database = os.getenv("AZURE_SQL_DATABASE")
        self.user = os.getenv("AZURE_SQL_USER")
        self.password = os.getenv("AZURE_SQL_PASSWORD")
        self.driver = '{ODBC Driver 18 for SQL Server}'

        self.sql_con_str = (
            f"Driver={self.driver};"
            f"Server=tcp:{self.server},1433;"
            f"Database={self.database};"
            f"Uid={self.user};"
            f"PWD={self.password};"
            f"Encrypt=yes;"
            f"TrustServerCertificate=no;"
            f"Connection Timeout=60;"
            f"Authentication=ActiveDirectoryPassword;"
        )
    
        # Connect  
        try:
            self.sql_conn = pyodbc.connect(self.sql_con_str)
            self.cursor = self.sql_conn.cursor()
            logging.info("Sql Connection is done.")
        except Exception as e:
            logging.error(f"❌ Failed to connect to Azure SQL: {e}")
            raise


    def insert_candidate_score(self, row: dict):
        insert_sql = """
        INSERT INTO dbo.Nam_Candidate_Scores (
            candidat, score_technique, detail_technique, justification_technique,
            score_rse, detail_rse, justification_rse,
            score_vente, detail_vente, justification_vente,
            fichier_source, total_score
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        self.cursor.execute(
            insert_sql,
            row["Candidat"],
            row["Score Technique /100"],
            row["Détail Technique (Exp/MC/Proj/Form)"],
            row["Justification Technique"],
            row["Score RSE /100"],
            row["Détail RSE (Exp/MC/Proj/Form)"],
            row["Justification RSE"],
            row["Score Vente /100"],
            row["Détail Vente (Exp/MC/Proj/Form)"],
            row["Justification Vente"],
            row["Fichier source"],
            row["Total Score"]
        )
        self.sql_conn.commit()
        logging.info(f"Inserted candidate score for {row['Candidat']}")

class CVProcessor:
    """Extract text and chunk CVs"""
    def __init__(self):
        self.endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
        self.deployment= os.getenv("AZURE_OPENAI_DEPLOYMENT")
        self.api_key = os.getenv("AZURE_OPENAI_API_KEY")
        self.model_name= os.getenv("AZURE_MODEL_NAME")
        
       
        self.client = OpenAI(
            api_key=self.api_key,
            base_url=self.endpoint
        )
    
        # ✅ Storage config
        conn = os.getenv("AzureWebJobsStorage")

        if not conn:
            raise ValueError("AZURE_STORAGE_CONNECTION_STRING not found")

        self.blob_service = BlobServiceClient.from_connection_string(conn)

        # Container where JSON files will go
        self.container_client = self.blob_service.get_container_client("cv-metadata")

        # Create container if not exists
        #try:
            #self.container_client.create_container()
        #except:
            #pass

    def extract_text_from_pdf_bytes(self, pdf_bytes: bytes) -> str:
        """Extract text directly from PDF bytes (Blob Storage)"""
        try:
            pdf_stream = io.BytesIO(pdf_bytes)
            pdf_reader = PyPDF2.PdfReader(pdf_stream)

            text = ""
            for page in pdf_reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"

            return text.strip()

        except Exception as e:
            logging.error(f"Error extracting PDF text: {e}")
            return ""
        
    def extract_enedis_metadata(self,text: str) -> dict:
        prompt = f"""
        
        You are an information extraction system.
        Your task is to extract structured information ONLY from the FIRST TWO PAGES of the provided document.
        
        Follow these instructions carefully:
        1. Identify the following sections in the document:
            - "Résidence"
            - "Syndic/Bailleur"
            - "Signatures et cachets"

            2. From each section, extract the required fields:
            
            === Résidence Section ===
            Extract:
            - Adresse

            === Syndic/Bailleur Section ===
            Extract:
            - Nom
            - Adresse
            - Interlocuteur
            - Tél
            - EMAIL

            === Sigantures et cachets ===
            from Prestataire extract:
            - adresse
            - Tél
            - Messagerie

            3. Rules:
            - Only extract information from the FIRST TWO PAGES.
            - If a field is missing, return null.
            - Do not hallucinate or guess values.
            - Clean and normalize extracted values (trim spaces, standardize phone/email).
            - Keep original language (French) when possible.

            Create a structured professional summary using exactely the folloxing format (do not change section titles):

            {{
                "syndic_bailleur": {
                    "nom": "",
                    "email": "",
                    "telephone": "",
                    "adresse": "",
                    "interlocuteur": ""
                },
                "residence": {
                    "adresse": "",
                },
                "prestataire": {
                    "adresse": "",
                    "telephone": "",
                    "messagerie": ""
                }
            }}

            file text:
            {text}
            """
        response = self.client.chat.completions.create(
            model=self.deployment,
            messages=[
                {"role": "system", "content": "You extract structured data and return JSON only."},
                {"role": "user", "content": prompt}
            ],
            temperature=0
        )

        try:
            return json.loads(response.choices[0].message.content.strip())
        except json.JSONDecodeError as e:
            logging.error(f"Invalid JSON returned by model: {e}")
            logging.error(response.choices[0].message.content)
            raise

       
    def extract_metadata(self,text: str) -> dict:
        prompt = f"""
            You are an information extraction system.

            From the CV text below create a professional candidate profile:
            - Do NOT invent information
            - Do NOT infer, assume, or improve anything.
            - Use ONLY what is explicitly written
            - If information is missing, write "Not specified"
            - Keep wording neutral and factual.

            Return ONLY a valid JSON object.
            Do not include markdown, code fences, or explanations.
            If a field is missing, return null or empty array.

            Follow EXACT schema: (do not change section titles):

            {{
                "candidate_title": "",
                "years_experience": "",
                "industry_domains": [],
                "technical_skills": [],
                "professional_experience": [
                    {{
                    "role": "",
                    "company": "",
                    "dates": "",
                    "description": ""
                    }}
                ],
                "education": [
                    {{
                    "degree": "",
                    "field": "",
                    "institution": ""
                    }}
                ],
                "languages": [],
                "additional_information": ""
                }}

            CV text:
            {text}
            """
        response = self.client.chat.completions.create(
            model=self.deployment,
            messages=[
                {"role": "system", "content": "You extract structured data and return JSON only."},
                {"role": "user", "content": prompt}
            ],
            temperature=0,
            response_format={ "type": "json_object" }   # 🔥 KEY FIX
        )

        try:
            return json.loads(response.choices[0].message.content.strip())
        except json.JSONDecodeError as e:
            logging.error(f"Invalid JSON returned by model: {e}")
            logging.error(response.choices[0].message.content)
            raise

   
    def save_metadata(self, candidate_name: str, metadata: dict):
        """Save metadata JSON to Blob Storage"""

        blob_name = f"{candidate_name}.json"
        blob_client = self.container_client.get_blob_client(blob_name)

        data = json.dumps(metadata, indent=2, ensure_ascii=False)

        blob_client.upload_blob(
            data,
            overwrite=True,
            content_type="application/json"
        )

        logging.info(f"Metadata uploaded: {blob_name}")

def transform_scoring_result(scoring_json: dict, source_filename: str):

    def build_detail_string(axe_data):
        blocs = axe_data["blocs"]

        return (
            f"Exp:{blocs['experience']['score']} | "
            f"MC:{blocs['mots_cles']['score']} | "
            f"Proj:{blocs['projets']['score']} | "
            f"Form:{blocs['formation']['score']}"
        )

    # Technique
    tech = scoring_json["axe_technique"]
    tech_score = tech["score_total"]
    tech_detail = build_detail_string(tech)
    tech_justification = tech["justification"]

    # RSE
    rse = scoring_json["axe_rse"]
    rse_score = rse["score_total"]
    rse_detail = build_detail_string(rse)
    rse_justification = rse["justification"]

    # Vente
    vente = scoring_json["axe_vente"]
    vente_score = vente["score_total"]
    vente_detail = build_detail_string(vente)
    vente_justification = vente["justification"]

    # Total Score (average of 3 axes)
    total_score = round((tech_score + rse_score + vente_score) / 3, 2)

    return {
        "Candidat": source_filename,
        "Score Technique /100": tech_score,
        "Détail Technique (Exp/MC/Proj/Form)": tech_detail,
        "Justification Technique": tech_justification,
        "Score RSE /100": rse_score,
        "Détail RSE (Exp/MC/Proj/Form)": rse_detail,
        "Justification RSE": rse_justification,
        "Score Vente /100": vente_score,
        "Détail Vente (Exp/MC/Proj/Form)": vente_detail,
        "Justification Vente": vente_justification,
        "Fichier source": source_filename,
        "Total Score": total_score
    }

def score_candidate(json_bytes: bytes):

    candidat_json = json.loads(json_bytes.decode("utf-8"))

    api_key = os.getenv("AZ_OPENAI_API_KEY")
    endpoint = os.getenv("AZ_OPENAI_ENDPOINT")
    deployment = os.getenv("AZ_OPENAI_DEPLOYMENT")
    api_version = os.getenv("AZ_OPENAI_API_VERSION")

    client = AzureOpenAI(
        api_key=api_key,
        azure_endpoint=endpoint,
        api_version=api_version
    )

    with open("system_prompt.txt", "r", encoding="utf-8") as f:
        system_prompt = f.read()
    #system_prompt = os.getenv("SCORING_SYSTEM_PROMPT")  # store your huge prompt in App Settings

    response = client.chat.completions.create(
        model=deployment,
        temperature=0.1,
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": f"""Voici le CV au format JSON d'un candidat.
            Analyse-le et score-le selon le barème fourni.

            ⚠️ IMPORTANT : Tu DOIS répondre avec EXACTEMENT ce format JSON (ne crée pas ton propre format) :
            {{
            "candidat": "Nom Prénom",
            "axe_technique": {{ "score_total": 0, "blocs": {{...}} }},
            "axe_rse": {{ "score_total": 0, "blocs": {{...}} }},
            "axe_vente": {{ "score_total": 0, "blocs": {{...}} }}
            }}

            CV du candidat :
            {json.dumps(candidat_json, ensure_ascii=False, indent=2)}"""}
            ])
    result = json.loads(response.choices[0].message.content)
    return result


class ExcelStorageClient:
    def __init__(self):
        conn_str = os.getenv("AzureWebJobsStorage")
        self.blob_service = BlobServiceClient.from_connection_string(conn_str)
        self.container_client = self.blob_service.get_container_client("cvjsons")

    def append_row(self, row: dict):
        blob_client = self.container_client.get_blob_client("Candidate_score.xlsx")

        try:
            # Download file
            excel_bytes = blob_client.download_blob().readall()
            wb = load_workbook(filename=BytesIO(excel_bytes))
            ws = wb.active
        except Exception:
            # If file doesn't exist → create new
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active

            # Add headers
            ws.append(list(row.keys()))

        # Append row
        ws.append(list(row.values()))

        # Save to memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Upload (overwrite)
        blob_client.upload_blob(output.getvalue(), overwrite=True)

        logging.info("Row appended to Excel file successfully")

app = func.FunctionApp()

@app.event_grid_trigger(arg_name="azeventgrid")
def EventGridTrigger(azeventgrid: func.EventGridEvent):
    logging.info('Python EventGrid trigger processed an event')
    event_data = azeventgrid.get_json()

    # Extract blob URL from Event Grid payload
    blob_url = event_data.get("url")
    if not blob_url:
        logging.warning("No blob URL in event data")
        return

    logging.info(f"New blob detected: {blob_url}")

    # Parse container name and blob name from URL
    parsed_url = urlparse(blob_url)
    path_parts = parsed_url.path.lstrip("/").split("/", 1)
    logging.info(f"path_parts: {path_parts}" )
    container_name = path_parts[0]
    blob_name = unquote(path_parts[1])
    logging.info(f"blobname: {blob_name}, containername: {container_name}")
     # Download PDF from storage
    conn_str = os.getenv("AzureWebJobsStorage")
    blob_service = BlobServiceClient.from_connection_string(conn_str)
    blob_client = blob_service.get_blob_client(container=container_name, blob=blob_name)

    if container_name == "cvfiles" and blob_name.lower().endswith(".pdf"):
        pdf_bytes = blob_client.download_blob().readall()
        logging.info(f"Downloaded {len(pdf_bytes)} bytes")

        # Process PDF
        processor = CVProcessor()
        candidate_name = Path(blob_name).stem

        text = processor.extract_text_from_pdf_bytes(pdf_bytes)
        logging.info("test are extracted")
        if not text:
            logging.warning(f"No text extracted from {blob_name}")
            return

        metadata = processor.extract_metadata(text)
        processor.save_metadata(candidate_name, metadata)
        logging.info(f"Processed CV '{candidate_name}' successfully")

    elif container_name == "cv-metadata" and blob_name.lower().endswith(".json"):
        logging.info("Processing new JSON metadata...")
        json_bytes = blob_client.download_blob().readall()
        
        processor = CVProcessor()
        candidate_name = Path(blob_name).stem

        scoring_result = score_candidate(json_bytes)
        logging.info("Scoring result:")
        formatted_row = transform_scoring_result(scoring_result,candidate_name)
        logging.info(json.dumps(formatted_row, indent=2, ensure_ascii=False))
        # Insert into Azure SQL
        try: 
            excel_client = ExcelStorageClient()
            excel_client.append_row(formatted_row)
        except Exception as e:
            logging.error(f"Failed to insert candidate score into exel file: {e}")
        try:
            sql_client = AzureSQLClient()  # connect to SQL
            sql_client.insert_candidate_score(formatted_row)  # insert the row
        except Exception as e:
            logging.error(f"Failed to insert candidate score into SQL: {e}")

    elif container_name == "cvjsons" and blob_name.lower().endswith(".json"):
       
        logging.info("Processing new JSON cvjsons...")
        json_bytes = blob_client.download_blob().readall()
        
        processor = CVProcessor()
        candidate_name = Path(blob_name).stem

        scoring_result = score_candidate(json_bytes)
        logging.info("Scoring result:")
        formatted_row = transform_scoring_result(scoring_result,candidate_name)
        logging.info(json.dumps(formatted_row, indent=2, ensure_ascii=False))
        try: 
            excel_client = ExcelStorageClient()
            excel_client.append_row(formatted_row)
        except Exception as e:
            logging.error(f"Failed to insert candidate score into excel: {e}")
    elif container_name == "enedis" and blob_name.lower().endswith(".pdf"):
        logging.info("Processing new Enedis files...")
        pdf_bytes = blob_client.download_blob().readall()
        logging.info(f"Downloaded {len(pdf_bytes)} bytes")
        # Process PDF
        processor = CVProcessor()
        enedis_name = Path(blob_name).stem

        text = processor.extract_text_from_pdf_bytes(pdf_bytes)
        logging.info(f"test are extracted {text}")
        if not text:
            logging.warning(f"No text extracted from {blob_name}")
            return
    else: 
        logging.info(f"no need to process anything")
